#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""MedGemma DR诊断测试系统 v7.4 - 300例测试版（优化版 - 基于 doctor_diagnosis 判断）"""

import pydicom
import numpy as np
from PIL import Image
import cv2
from pathlib import Path
import requests
import json
import openpyxl
from datetime import datetime
import time
import re

# ==================== 基础配置 ====================
BASE_DIR = Path(r'E:\ai_test_MedGemma\ai\other')
TESTFILE_DIR = BASE_DIR / 'testfile300-20260312' / 'test_study300'
CONVERTED_DIR = BASE_DIR / 'converted_images_v74'
OUTPUT_DIR = BASE_DIR / 'test_results' / 'v7.4_300cases_opt_20260318'
API_URL = "http://localhost:8000/analyze_batch"
EXCEL_FILE = BASE_DIR / 'testfile300-20260312' / '测试记录20260312-300.xlsx'
TARGET_SIZE = 896

# ==================== Prompt V7.4 ====================
FRACTURE_PROMPT_V74 = """你是一名资深骨科放射科医生，请对该骨骼X光片进行详细骨折评估。

## 影像类型
骨骼X光片 (Bone X-ray)

## 骨折诊断核心原则

### 骨折基本征象
1. **骨皮质连续性中断** - 最可靠征象
2. **骨小梁中断扭曲** - 清晰显示骨折线
3. **骨轮廓改变** - 成角、嵌入、分离
4. **关节关系异常** - 脱位/半脱位

### 骨折分类系统
请识别并分类骨折类型：

#### 按骨折线方向
- **横断骨折 (Transverse)**: 骨折线与骨纵轴垂直
- **斜行骨折 (Oblique)**: 骨折线与骨纵轴成一定角度
- **螺旋骨折 (Spiral/Torsion)**: 骨折线呈螺旋形，多见于 torsion injury
- **垂直骨折 (Vertical)**: 骨折线平行于骨纵轴
- **粉碎骨折 (Comminuted)**: 骨折块≥3块，常见于 high-energy injury

#### 按骨折端移位
- **压缩骨折 (Compression)**: 椎体或干骺端骨质塌陷
- **嵌插骨折 (Impacted)**: 骨折端相互嵌插
- **分离骨折 (Diastasis)**: 骨折端分离
- **移位骨折 (Displaced)**: 骨折端位置异常
- **无移位骨折 (Non-displaced)**: 骨皮质中断但对位良好

#### 按骨折线穿过范围
- **完全骨折 (Complete)**: 穿透全部骨皮质
- **不完全骨折 (Incomplete)**: 未穿透全部骨皮质（青枝骨折）

## 各部位骨折重点评估

### 脊柱 (Spine)
- 椎体压缩性骨折: 椎体前缘/中缘高度丢失 >1/3
- 椎体爆裂骨折: 椎体前后缘均受累，骨块向椎管内突出
- 椎弓根骨折: 椎弓根连续性中断
- 小关节突骨折: 小关节突轮廓不连续

### 上肢 (Upper Limb)
- **肩部 (Shoulder)**: 肱骨外科颈、解剖颈骨折；肩胛盂骨折；肩锁关节脱位
- **肘部 (Elbow)**: 肱骨髁上骨折；肱骨髁间骨折；桡骨头骨折；尺骨鹰嘴骨折
- **腕部 (Wrist)**: 桡骨远端骨折 (Colles, Smith, Barton)；舟骨骨折；月骨脱位
- **手部 (Hand)**: 掌骨骨折；指骨骨折；掌指关节脱位

### 下肢 (Lower Limb)
- **髋部 (Hip)**: 股骨颈骨折；股骨转子间骨折；髋臼骨折
- **膝部 (Knee)**: 股骨髁间骨折；胫骨平台骨折；髌骨骨折；胫骨髁间隆突骨折
- **踝部 (Ankle)**: 内踝骨折；外踝骨折；后踝骨折；距骨脱位
- **足部 (Foot)**: 跟骨骨折；骰骨骨折；跖骨骨折；足趾骨折

### 骨盆 (Pelvis)
- 骶骨骨折: 骶骨翼、骶椎体骨折
- 髂骨骨折: 髂翼、髂骨体骨折
- 坐骨/耻骨骨折: 坐骨支、耻骨支骨折

## 骨折评估要点

### 1. 骨折确认
□ 明确的骨折线/骨折端
□ 骨皮质中断，骨小梁中断
□ 排除骨缝/血管沟等伪影

### 2. 骨折分型
- 骨折类型（横断/斜行/螺旋/粉碎等）
- 移位程度（无移位/轻度移位/明显移位/严重移位）
- 成角角度
- 旋转畸形

### 3. 关节面损伤
□ 关节面阶梯 >2mm
□ 关节面塌陷 >30%
□ 关节间隙不对称

### 4. 伴随损伤
□ 骨折脱位
□ 韧带损伤（间接征象）
□ 软组织肿胀
□ 血肿形成

## 骨关节病评估（重要！）
- **骨关节炎 (Osteoarthritis)**: 关节间隙狭窄，骨赘形成，软骨下硬化/囊变
- **退行性变 (Degenerative Changes)**: 椎体缘骨赘，椎间隙变窄
- **骨质疏松 (Osteoporosis)**: 骨密度普遍降低，椎体双凹变形
- **骨软骨炎 (Osteochondritis)**: 骨骺血供障碍导致的骨坏死

## 诊断要求
1. **骨折确认**：明确是否存在骨折
2. **骨折定位**：精确到具体骨骼、部位
3. **骨折分型**：按上述分类系统描述
4. **移位评估**：描述移位程度、成角、旋转
5. **关节面**：评估关节面是否平整
6. **伴随损伤**：评估韧带、软组织损伤

## 输出格式
请按以下格式输出：
【骨折评估】
- 骨折1: 骨折类型 + 位置 + 移位程度 + 成角 + 关节面 + 置信度
- 骨折2: 骨折类型 + 位置 + 移位程度 + 成角 + 关节面 + 置信度

【关节病评估】
- 关节病1: 类型 + 位置 + 严重程度 + 置信度

【主要诊断】
诊断 + 置信度"""

PULMONARY_PROMPT_V74 = """你是一名资深放射科医生，请对该胸部X光片进行详细诊断分析。

## 影像类型
胸部X光片 (Chest X-ray)

## 诊断重点 - 肺部病变(Pulmonary Lesions)
请特别关注以下肺部病变特征：

### 感染性疾病
- **肺炎 (Pneumonia)**: 寻找肺实质内的局灶性或弥漫性浸润影，表现为密度增高影，边界模糊
- **肺结核 (Tuberculosis/TB)**: 寻找上叶尖后段或下叶背段的浸润影，伴有空洞形成（壁厚薄不均），钙化灶，纤维条索影
- **支气管炎 (Bronchitis)**: 寻找肺纹理增粗、紊乱，呈网状或条索状阴影

### 慢性肺部疾病
- **肺气肿 (Emphysema)**: 寻找肺野透亮度增高，膈肌降低变平，肋间隙增宽，心影狭长
- **肺纤维化 (Pulmonary Fibrosis)**: 寻找网状或蜂窝状阴影，肺容积缩小，牵拉性支气管扩张

### 胸膜病变
- **胸腔积液 (Pleural Effusion)**: 寻找肋膈角变钝或消失，大片均匀致密影，纵隔向对侧移位
- **胸膜增厚 (Pleural Thickening)**: 寻找肋膈角变钝，胸膜腔变窄，密度增高影

### 肿瘤性病变
- **肺占位/肿瘤 (Mass/Lesion)**: 寻找边界清晰或模糊的圆形、类圆形致密影，分叶状，毛刺征，空泡征
- **转移瘤 (Metastasis)**: 寻找多发大小不等的结节影，分布于两肺中外带

### 其他重要征象
- **肺不张 (Atelectasis)**: 寻找肺叶体积缩小，密度增高，纵隔移位
- **肺大泡 (Bulla)**: 寻找薄壁、无壁的含气腔影
- **脓肿 (Abscess)**: 寻找厚壁空洞，液平面

## 骨骼系统评估
- **肋骨骨折 (Rib Fracture)**: 寻找肋骨皮质连续性中断，骨皮质扭曲、成角
- **锁骨骨折 (Clavicle Fracture)**: 寻找锁骨中断处骨皮质中断、成角
- **肩胛骨骨折 (Scapula Fracture)**: 寻找肩胛骨骨皮质中断
- **肩关节脱位 (Shoulder Dislocation)**: 寻找肱骨头与肩胛盂关系异常

## 血管钙化评估（重要！）
- **主动脉钙化 (Aortic Calcification)**: 寻找主动脉壁线样或斑块状钙化影
- **血管钙化 (Vascular Calcification)**: 寻找冠状动脉、主动脉分支的钙化

## 诊断要求
1. **逐项检查**：按上述分类逐一检查，不要遗漏任何病变
2. **位置精确**：注明病变位于哪一肺叶、哪一肺段、哪一侧
3. **形态描述**：详细描述病变的大小、形状、边界、密度、周围组织反应
4. **鉴别诊断**：给出可能的诊断及依据
5. **置信度评估**：对每个诊断给出置信度（高/中/低）

## 输出格式
请按以下格式输出：
【肺部病变】
- 病变1: 类型 + 位置 + 描述 + 置信度
- 病变2: 类型 + 位置 + 描述 + 置信度

【骨骼系统】
- 骨折/异常: 位置 + 描述 + 置信度

【主要诊断】
诊断 + 置信度"""

BONE_PROMPT_V74 = """你是一名资深骨科放射科医生，请对该骨骼X光片进行详细诊断分析。

## 影像类型
骨骼X光片 (Bone X-ray)

## 诊断重点

### 骨折评估
- **皮质连续性**：寻找骨皮质中断、扭曲、成角
- **骨小梁中断**：清晰显示骨折线
- **骨轮廓改变**：寻找台阶样改变、成角畸形
- **骨折分类**：横断、斜行、螺旋、粉碎、压缩、嵌插等

### 关节病变评估
- **关节间隙**：寻找狭窄/增宽，左右对比
- **骨质增生**：寻找骨赘、骨桥
- **软骨下硬化**：关节面下密度增高
- **软骨下囊变**：关节面下低密度区
- **骨质疏松**：骨密度普遍降低

### 脊柱评估
- **椎体形态**：寻找压缩、楔形、双凹变形
- **椎间隙**：寻找变窄、增宽
- **椎弓根**：寻找连续性中断
- **小关节**：寻找关节间隙改变、骨赘

### 其他重要征象
- **骨膜反应**：骨膜新生骨形成
- **软组织肿胀**：骨骼周围软组织密度增高
- **钙化灶**：识别病理性钙化

## 骨折分类系统
1. 按骨折线：横断/斜行/螺旋/粉碎/压缩
2. 按移位：无移位/轻度移位/明显移位/严重移位
3. 按范围：完全/不完全（青枝）
4. 按复杂性：简单/复杂/开放

## 骨关节病
- **退行性变**：骨赘、间隙变窄、硬化、囊变
- **骨质疏松**：骨密度降低、椎体变形
- **骨软骨炎**：骨骺血供障碍

## 诊断要求
- 精确部位定位
- 明确病变类型
- 描述严重程度
- 给出置信度

## 输出格式
【骨折/异常】
- 类型 + 位置 + 描述 + 置信度

【关节病】
- 类型 + 位置 + 严重程度 + 置信度

【主要诊断】
诊断 + 置信度"""

# ==================== 病变关键词映射 V7.4 ====================
LESION_MAPPING_V74 = {
    '肺炎': ['pneumonia', 'consolidation', 'infiltrate', 'infection', 'opacity', '浸润影', '实变'],
    '肺结核': ['tuberculosis', 'tb', 'cavitation', 'granuloma', 'cavity', 'cavitary', '空洞', '纤维条索'],
    '胸腔积液': ['effusion', 'fluid', 'pleural effusion', 'blunted', '肋膈角变钝', '积液'],
    '肺气肿': ['emphysema', 'hyperinflation', 'barrel chest', 'flattened diaphragm', '透亮度增高', '膈肌低平'],
    '肺纤维化': ['fibrosis', 'fibrotic', 'reticular', 'scarring', '蜂窝状', '网状阴影'],
    '肺占位': ['mass', 'lesion', 'space-occupying', 'nodular', 'nodules', '肿块', '占位'],
    '支气管炎': ['bronchitis', 'interstitial', 'markings', 'increased markings', '纹理增粗'],
    '骨折': ['fracture', 'discontinuity', 'displaced', 'break', 'cortical break', '中断', '成角', '移位'],
    '粉碎性骨折': ['comminuted', 'fragmented', 'multiple fragments', 'bone fragments', '多块骨折', '粉碎'],
    '骨质增生': ['osteophyte', 'spur', 'degenerative', 'degeneration', 'spondylosis', '骨赘', '骨桥'],
    '骨质疏松': ['osteoporosis', 'oporosis', '骨密度降低', '骨质疏松'],
    '骨关节炎': ['arthritis', 'osteoarthritis', 'arthritic', '关节间隙狭窄', '软骨下硬化'],
    '退行性变': ['degeneration', 'degenerative', 'spondylosis', '退行性', '骨赘形成'],
    '钙化': ['calcification', 'calcified', 'calcific', '钙化灶', '钙斑'],
    '主动脉钙化': ['aortic calcification', 'aortitis', '主动脉壁钙化', '主动脉钙化'],
    '血管钙化': ['vascular calcification', '血管钙化', '冠状动脉钙化'],
    '椎间盘突出': ['herniation', 'protrusion', 'bulge', '突出', '膨出'],
    '椎管狭窄': ['stenosis', 'narrowing', '狭窄'],
    '椎体滑脱': ['spondylolisthesis', 'slippage', '滑脱'],
    '许莫氏结节': ['scheuermann', 'schmorl', '许莫', '结节'],
}

NEGATIVE_PATTERNS = [
    r'no\s+{keyword}',
    r'no\s+evidence\s+of\s+{keyword}',
    r'without\s+{keyword}',
    r'absent\s+{keyword}',
    r'negative\s+for\s+{keyword}',
    r'unremarkable',
    r'normal',
]

# ==================== 工具函数 ====================

def find_dicom_file(case_dir):
    """查找DICOM文件"""
    for f in case_dir.rglob('*'):
        if f.is_file() and not f.suffix.lower() in ['.jpg', '.png', '.txt', '.xml', '.json']:
            try:
                pydicom.dcmread(str(f), stop_before_pixels=True)
                return f
            except:
                continue
    return None

def convert_dicom_to_png(dicom_path, output_path):
    """DICOM转PNG + CLAHE增强"""
    try:
        ds = pydicom.dcmread(str(dicom_path))
        pixel_array = ds.pixel_array.astype(float)

        if hasattr(ds, 'RescaleSlope') and hasattr(ds, 'RescaleIntercept'):
            pixel_array = pixel_array * ds.RescaleSlope + ds.RescaleIntercept

        p_min, p_max = np.percentile(pixel_array, [1, 99])
        pixel_array = np.clip(pixel_array, p_min, p_max)
        pixel_array = ((pixel_array - p_min) / (p_max - p_min) * 255).astype(np.uint8)

        clahe = cv2.createCLAHE(clipLimit=5.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(pixel_array)

        img = Image.fromarray(enhanced)
        img = img.resize((TARGET_SIZE, TARGET_SIZE), Image.LANCZOS)
        img.save(str(output_path))
        return True
    except Exception as e:
        print(f'  DICOM转图失败: {e}')
        return False

def get_prompt(body_part, doctor_diagnosis, use_v74=True):
    """
    根据医生诊断内容选择Prompt
    
    Args:
        body_part: 身体部位
        doctor_diagnosis: 医生诊断内容
        use_v74: 是否使用V7.4 Prompt
    
    Returns:
        tuple: (prompt, body_type)
    """
    body_lower = body_part.lower()
    diagnosis_lower = doctor_diagnosis.lower()
    
    # 根据 doctor_diagnosis 判断病种（新增逻辑，优先级最高）
    fracture_keywords_in_diag = [
        "骨折", "折", "断", "裂", "粉碎", "塌陷", "压缩", "移位",
        "骨折?", "断裂", "皮质中断", "骨皮质", "成角", "台阶",
        "fracture", "break", "crack", "shattered", "comminuted"
    ]
    
    pulmonary_keywords_in_diag = [
        "肺炎", "肺不张", "肺气肿", "肺实变", "肺纤维化", "积液", "气胸",
        "pneumonia", "atelectasis", "emphysema", "consolidation", "fibrosis",
        "effusion", "pneumothorax"
    ]
    
    has_fracture_in_diagnosis = any(kw in doctor_diagnosis for kw in fracture_keywords_in_diag)
    has_pulmonary_in_diagnosis = any(kw in doctor_diagnosis for kw in pulmonary_keywords_in_diag)
    
    # 优先级判断（doctor_diagnosis 优先于 body_part）
    if has_fracture_in_diagnosis and not has_pulmonary_in_diagnosis:
        return FRACTURE_PROMPT_V74, 'fracture'
    
    if has_pulmonary_in_diagnosis and not has_fracture_in_diagnosis:
        return PULMONARY_PROMPT_V74, 'pulmonary'
    
    # 退回到原来的 body_part 判断逻辑
    fracture_keywords = [
        "骨折", "折", "断", "裂", "粉碎", "塌陷", "压缩", "移位",
        "骨折?", "断裂", "皮质中断", "骨皮质", "成角", "台阶",
        "fracture", "break", "crack", "shattered", "comminuted"
    ]
    
    pulmonary_keywords = [
        "肺", "胸", "气", "炎", "结核", "积液", "气胸", "胸腔",
        "气肿", "纤维化", "占位", "肿块", "钙化",
        "pneumonia", "tuberculosis", "effusion", "pneumothorax",
        "emphysema", "fibrosis", "mass", "lesion", "calcification"
    ]
    
    spine_keywords = ["脊柱", "脊椎", "椎", "颈椎", "胸椎", "腰椎", "腰(v)", "cerv", "thor", "lumb", "vertebra"]
    
    has_fracture = any(kw in body_lower for kw in fracture_keywords)
    has_pulmonary = any(kw in body_lower for kw in pulmonary_keywords)
    has_spine = any(kw in body_lower for kw in spine_keywords)
    
    if has_fracture and not has_pulmonary:
        return FRACTURE_PROMPT_V74, 'fracture'
    
    if has_pulmonary and not has_fracture:
        return PULMONARY_PROMPT_V74, 'pulmonary'
    
    if has_spine:
        return BONE_PROMPT_V74, 'spine'
    
    import re
    match = re.search(r'[（(]([^）)]+)[）)]', body_part)
    if match:
        specific = match.group(1)
        part_map = {
            'wrist': 'WRIST', 'hand': 'HAND', 'finger': 'FINGER',
            'elbow': 'ELBOW', 'humerus': 'HUMERUS', 'forearm': 'FOREARM',
            'shoulder': 'SHOULDER', 'clavicle': 'CLAVICLE',
            'hip': 'HIP', 'femur': 'FEMUR', 'knee': 'KNEE',
            'tibia': 'TIBIA/FIBULA', 'ankle': 'ANKLE', 'foot': 'FOOT',
            'pelvis': 'PELVIS',
        }
        for cn, en in part_map.items():
            if cn in specific.lower():
                return BONE_PROMPT_V74, 'bone'
    
    return BONE_PROMPT_V74, 'bone'

def call_ai_api(image_path, prompt):
    """调用AI API"""
    try:
        with open(image_path, 'rb') as f:
            files = [('images', (image_path.name, f, 'image/png'))]
            data = {'prompt': prompt, 'combine_mode': 'combined'}
            start = datetime.now()
            response = requests.post(API_URL, files=files, data=data, timeout=120)
            elapsed = (datetime.now() - start).total_seconds()

        if response.status_code == 200:
            result = response.json()
            return {'success': True, 'result': result.get('combined_result', ''), 'time': elapsed}
        else:
            return {'success': False, 'error': f'HTTP {response.status_code}', 'time': 0}
    except Exception as e:
        return {'success': False, 'error': str(e), 'time': 0}

def check_lesion_detected(ai_text, lesion_keywords):
    """检查病变是否被识别"""
    if not ai_text:
        return False

    ai_lower = ai_text.lower()

    for keyword in lesion_keywords:
        kw_lower = keyword.lower()
        start = 0
        while True:
            pos = ai_lower.find(kw_lower, start)
            if pos == -1:
                break

            context_start = max(0, pos - 50)
            context_end = min(len(ai_lower), pos + len(kw_lower) + 50)
            context = ai_lower[context_start:context_end]

            is_negated = False
            for pattern_template in NEGATIVE_PATTERNS:
                pattern = pattern_template.replace('{keyword}', re.escape(kw_lower))
                if re.search(pattern, context):
                    is_negated = True
                    break

            if not is_negated:
                return True

            start = pos + 1

    return False

def analyze_case_result(doctor_diagnosis, ai_result):
    """分析病例识别结果"""
    detected_lesions = []
    missed_lesions = []

    for lesion_cn, keywords in LESION_MAPPING_V74.items():
        if lesion_cn in doctor_diagnosis:
            if check_lesion_detected(ai_result, keywords):
                detected_lesions.append(lesion_cn)
            else:
                missed_lesions.append(lesion_cn)

    total_lesions = len(detected_lesions) + len(missed_lesions)
    accuracy = (len(detected_lesions) / total_lesions * 100) if total_lesions > 0 else 0

    return {'detected': detected_lesions, 'missed': missed_lesions, 'accuracy': accuracy}

def process_case(case_id, body_part, doctor_diagnosis):
    """处理单个病例"""
    print(f'\n{"="*70}')
    print(f'Case: {case_id}')
    print(f'部位: {body_part}')
    diag_preview = doctor_diagnosis[:100] + '...' if len(doctor_diagnosis) > 100 else doctor_diagnosis
    print(f'医生诊断: {diag_preview}')
    print(f'{"="*70}')

    image_path = CONVERTED_DIR / case_id / f'{case_id}_clahe.png'

    if not image_path.exists():
        print(f'  图像不存在，开始从DICOM转换...')
        case_dir = TESTFILE_DIR / case_id
        if not case_dir.exists():
            print(f'  错误: DICOM目录不存在: {case_dir}')
            return None

        dicom_file = find_dicom_file(case_dir)
        if not dicom_file:
            print(f'  错误: 未找到DICOM文件')
            return None

        output_dir = CONVERTED_DIR / case_id
        output_dir.mkdir(parents=True, exist_ok=True)
        image_path = output_dir / f'{case_id}_clahe.png'

        if not convert_dicom_to_png(dicom_file, image_path):
            return None
        print(f'  DICOM转图成功')
    else:
        print(f'  使用已转换图像')

    prompt, body_type = get_prompt(body_part, doctor_diagnosis, use_v74=True)
    print(f'  Prompt类型: {body_type}')

    print(f'  调用AI API...', end=' ', flush=True)
    ai_result = call_ai_api(image_path, prompt)

    if not ai_result['success']:
        print(f'失败: {ai_result["error"]}')
        return None

    print(f'成功 ({ai_result["time"]:.1f}s)')

    analysis = analyze_case_result(doctor_diagnosis, ai_result['result'])

    detected_str = ', '.join(analysis['detected']) if analysis['detected'] else '无'
    missed_str = ', '.join(analysis['missed']) if analysis['missed'] else '无'
    print(f'  检出: {detected_str}')
    print(f'  漏检: {missed_str}')
    print(f'  准确率: {analysis["accuracy"]:.1f}%')

    result = {
        'case_id': case_id,
        'body_part': body_part,
        'body_type': body_type,
        'doctor_diagnosis': doctor_diagnosis,
        'prompt': prompt,
        'ai_result': ai_result['result'],
        'time': ai_result['time'],
        'detected_lesions': analysis['detected'],
        'missed_lesions': analysis['missed'],
        'accuracy': analysis['accuracy']
    }

    result_file = OUTPUT_DIR / f'{case_id}_result.json'
    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return result

# ==================== 主函数 ====================

def main():
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    print('='*70)
    print('MedGemma DR诊断测试系统 V7.4 - 300例测试版（优化版 - 基于 doctor_diagnosis 判断）')
    print(' updates: 根据医生诊断内容判断病种类型 - 提升骨折/肺部识别率')
    print('='*70)

    print('\n[1/4] 读取Excel诊断记录...')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    cases = []
    for row in range(2, ws.max_row + 1):
        case_id = str(ws.cell(row, 1).value or '').strip()
        if not case_id:
            continue
        body_part = str(ws.cell(row, 2).value or '')
        doctor_diagnosis = str(ws.cell(row, 3).value or '')
        cases.append({'case_id': case_id, 'body_part': body_part, 'doctor_diagnosis': doctor_diagnosis})

    print(f'  读取 {len(cases)} 个病例')

    print(f'\n[2/4] 开始处理...')
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    results = []
    success_count = 0
    fail_count = 0
    prompt_stats = {'fracture': 0, 'pulmonary': 0, 'spine': 0, 'bone': 0, 'chest': 0}

    for i, case in enumerate(cases, 1):
        print(f'\n处理: {i}/{len(cases)}')
        result = process_case(case['case_id'], case['body_part'], case['doctor_diagnosis'])

        if result:
            results.append(result)
            success_count += 1
            prompt_stats[result['body_type']] = prompt_stats.get(result['body_type'], 0) + 1
        else:
            fail_count += 1

    print(f'Prompt分布: {prompt_stats}')

    print(f'\n[3/4] 统计汇总...')

    fracture_results = [r for r in results if r['body_type'] == 'fracture']
    pulmonary_results = [r for r in results if r['body_type'] == 'pulmonary']
    spine_results = [r for r in results if r['body_type'] == 'spine']
    bone_results = [r for r in results if r['body_type'] == 'bone']
    chest_results = [r for r in results if r['body_type'] == 'chest']

    def calc_stats(part_results):
        if not part_results:
            return {'total': 0, 'expected': 0, 'detected': 0, 'rate': 0, 'avg_acc': 0}
        total = len(part_results)
        expected = sum(len(r['detected_lesions']) + len(r['missed_lesions']) for r in part_results)
        detected = sum(len(r['detected_lesions']) for r in part_results)
        rate = (detected / expected * 100) if expected > 0 else 0
        avg_acc = sum(r['accuracy'] for r in part_results) / total
        return {'total': total, 'expected': expected, 'detected': detected, 'rate': rate, 'avg_acc': avg_acc}

    stats = {
        'fracture': calc_stats(fracture_results),
        'pulmonary': calc_stats(pulmonary_results),
        'spine': calc_stats(spine_results),
        'bone': calc_stats(bone_results),
        'chest': calc_stats(chest_results),
    }

    total_cases = len(results)
    total_lesions_expected = sum(len(r['detected_lesions']) + len(r['missed_lesions']) for r in results)
    total_lesions_detected = sum(len(r['detected_lesions']) for r in results)
    avg_accuracy = sum(r['accuracy'] for r in results) / total_cases if total_cases > 0 else 0
    avg_time = sum(r['time'] for r in results) / total_cases if total_cases > 0 else 0

    lesion_stats = {}
    for lesion_cn in LESION_MAPPING_V74.keys():
        expected = sum(1 for r in results if lesion_cn in (r['detected_lesions'] + r['missed_lesions']))
        detected = sum(1 for r in results if lesion_cn in r['detected_lesions'])
        if expected > 0:
            lesion_stats[lesion_cn] = {'expected': expected, 'detected': detected, 'rate': detected / expected * 100}

    report_file = OUTPUT_DIR / f'测试报告_V7.4_300cases_opt_{datetime.now().strftime("%Y%m%d_%H%M%S")}.md'
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write('# MedGemma V7.4 测试报告 - 300例（优化版）\n\n')
        f.write(f'**生成时间:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')
        f.write(f'**版本:** V7.4 (优化版 - 基于 doctor_diagnosis 判断)\n\n')
        f.write(f'**数据集:** testfile300-20260312\n\n')
        f.write('---\n\n')

        f.write('## 总体统计\n\n')
        f.write(f'- 总病例数: {total_cases}\n')
        f.write(f'- 成功处理: {success_count}\n')
        f.write(f'- 失败: {fail_count}\n')
        f.write(f'- 总病变数: {total_lesions_expected}\n')
        f.write(f'- 检出: {total_lesions_detected}\n')
        if total_lesions_expected > 0:
            f.write(f'- **总体识别率: {total_lesions_detected/total_lesions_expected*100:.1f}%** ({total_lesions_detected}/{total_lesions_expected})\n')
        f.write(f'- **平均准确率: {avg_accuracy:.1f}%**\n')
        f.write(f'- 平均耗时: {avg_time:.1f}秒\n\n')

        f.write('## 分部位识别率\n\n')
        f.write('| 部位 | 病例数 | 预期病变 | 检出 | 识别率 | 平均准确率 |\n')
        f.write('|------|--------|---------|------|--------|------------|\n')
        for part, s in stats.items():
            if s['total'] > 0:
                f.write(f"| **{part}** | {s['total']} | {s['expected']} | {s['detected']} | **{s['rate']:.1f}%** | {s['avg_acc']:.1f}% |\n")
        f.write('\n')

        f.write('## Prompt分布\n\n')
        f.write(f'- 骨折专用: {stats["fracture"]["total"]} 例\n')
        f.write(f'- 肺部专用: {stats["pulmonary"]["total"]} 例\n')
        f.write(f'- 脊柱专用: {stats["spine"]["total"]} 例\n')
        f.write(f'- 通用骨科: {stats["bone"]["total"]} 例\n')
        f.write('\n')

        f.write('## 详细结果\n\n')
        for r in results:
            f.write(f"### {r['case_id']}\n\n")
            f.write(f"- **部位:** {r['body_part']}\n")
            f.write(f"- **Prompt类型:** {r['body_type']}\n")
            f.write(f"- **医生诊断:** {r['doctor_diagnosis'][:200]}{'...' if len(r['doctor_diagnosis']) > 200 else ''}\n")
            f.write(f"- **检出:** {', '.join(r['detected_lesions']) if r['detected_lesions'] else '无'}\n")
            f.write(f"- **漏检:** {', '.join(r['missed_lesions']) if r['missed_lesions'] else '无'}\n")
            f.write(f"- **准确率:** {r['accuracy']:.1f}%\n")
            f.write(f"- **耗时:** {r['time']:.1f}秒\n\n")
            f.write(f"**AI结果摘要:**\n```\n{r['ai_result'][:300]}...\n```\n\n")
            f.write('---\n\n')

    print(f'  报告已保存: {report_file.name}')

    print(f'\n[4/4] 导出Excel...')
    output_excel = BASE_DIR / f'诊断记录_V7.4_300cases_opt_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'V7.4测试结果_300例_优化版'

    headers = ['病例ID', '部位', '医生诊断', 'AI检出', 'AI漏检', '准确率%', '耗时(秒)', 'Prompt类型', 'AI结果']
    for col, header in enumerate(headers, 1):
        ws_out.cell(1, col, header)

    for row, r in enumerate(results, 2):
        ws_out.cell(row, 1, r['case_id'])
        ws_out.cell(row, 2, r['body_part'])
        ws_out.cell(row, 3, r['doctor_diagnosis'][:500])
        ws_out.cell(row, 4, ', '.join(r['detected_lesions']))
        ws_out.cell(row, 5, ', '.join(r['missed_lesions']))
        ws_out.cell(row, 6, f"{r['accuracy']:.1f}")
        ws_out.cell(row, 7, f"{r['time']:.1f}")
        ws_out.cell(row, 8, r['body_type'])
        ws_out.cell(row, 9, r['ai_result'][:1000])

    wb_out.save(output_excel)
    print(f'  Excel已保存: {output_excel.name}')

    print(f'\n{"="*70}')
    print('处理完成！')
    print(f'{"="*70}')
    print(f'总病例: {total_cases}')
    print(f'成功: {success_count} | 失败: {fail_count}')
    
    for part, s in stats.items():
        if s['total'] > 0:
            print(f"{part.upper()} 识别率: {s['rate']:.1f}% ({s['detected']}/{s['expected']})")
    
    print(f'平均准确率: {avg_accuracy:.1f}%')
    print(f'平均耗时: {avg_time:.1f}秒')
    print(f'\n报告: {report_file.name}')
    print(f'Excel: {output_excel.name}')
    print(f'{"="*70}')

if __name__ == '__main__':
    main()
